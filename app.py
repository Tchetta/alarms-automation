import os
# 1. Stability Layer: Prevent OpenBLAS memory allocation crashes in multi-threaded runtimes
os.environ["OPENBLAS_NUM_THREADS"] = "1"

import zipfile
import io
import datetime
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Set browser page layout properties
st.set_page_config(
    page_title="Power Alarms",
    page_icon="⚡",
    layout="centered"
)

st.title("⚡ Power Alarms")
st.markdown(
    "Upload your raw alarm file (supports raw **.xlsx** or compressed **.zip**). "
    "The dashboard will automatically handle decompression and reference correlation maps."
)

# Step 1: Connect and validate the local database reference file (ref1.xlsx)
if not os.path.exists("ref1.xlsx"):
    st.error("❌ Could not find 'ref1.xlsx' in the hosting directory layout. Verify its placement.")
    st.stop()

@st.cache_data
def load_reference_data():
    try:
        # Read reference file using pandas
        df_ref = pd.read_excel("ref1.xlsx")
        
        # Strip whitespaces from string column headers
        df_ref.columns = [str(c).strip() for c in df_ref.columns]
        
        required_ref_cols = ['Site ID', 'New Power Owner', 'Children', 'Backbone Site']
        for col in required_ref_cols:
            if col not in df_ref.columns:
                st.error(f"❌ Reference schema structure error! Missing column header: '{col}'")
                st.stop()
        
        # Clean up IDs and set index for ultra-fast dictionary lookups
        df_ref['Site ID'] = pd.to_numeric(df_ref['Site ID'], errors='coerce')
        df_ref = df_ref.dropna(subset=['Site ID'])
        df_ref['Site ID'] = df_ref['Site ID'].astype(int)
        
        # Build lookup maps mirroring the JS engine
        power_co_map = df_ref.set_index('Site ID')['New Power Owner'].fillna('IHS').to_dict()
        children_map = df_ref.set_index('Site ID')['Children'].fillna(0).to_dict()
        backbone_map = df_ref.set_index('Site ID')['Backbone Site'].fillna('').astype(str).str.strip().str.lower().to_dict()
        
        return power_co_map, children_map, backbone_map
    except Exception as e:
        st.error(f"❌ Error compiling reference dictionaries: {e}")
        st.stop()

power_co_map, children_map, backbone_map = load_reference_data()

# Step 2: File Upload Interface Component
uploaded_file = st.file_uploader(
    "Choose a Power Alarm Export File", 
    type=["xlsx", "zip"], 
    label_visibility="collapsed"
)

if uploaded_file is not None:
    # Processing block trigger
    if st.button("🚀 Generate Formatted Alarm Report", use_container_width=True):
        status_log = st.empty()
        
        try:
            file_bytes = uploaded_file.read()
            alarm_df = None
            
            # Auto-Detect and handle ZIP archive extractions
            if uploaded_file.name.lower().endswith('.zip'):
                status_log.info("ℹ️ Zip container detected! Initiating micro-decompression extraction layers...")
                with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
                    excel_files = [f for f in z.namelist() if f.lower().endswith('.xlsx') and not f.startswith('__MACOSX')]
                    if not excel_files:
                        raise ValueError("Invalid Archive structure! No valid .xlsx spreadsheets found inside the .zip container.")
                    
                    status_log.info(f"ℹ️ Extracting internal file target: \"{excel_files[0]}\"...")
                    with z.open(excel_files[0]) as f:
                        alarm_df = pd.read_excel(f)
            else:
                status_log.info("ℹ️ Reading uploaded data payload spreadsheet...")
                alarm_df = pd.read_excel(io.BytesIO(file_bytes))
            
            # Clean up headers
            alarm_df.columns = [str(c).strip() for c in alarm_df.columns]
            
            req_cols = ['Site ID', 'Site Name', 'Ticket ID', 'Alarm Name', 'First Occurred On', 'Duration(hh:mm:ss)', 'Last Occurred On']
            missing_cols = [col for col in req_cols if col not in alarm_df.columns]
            if missing_cols:
                raise ValueError(f"Data structure error! Upload missing columns: {', '.join(missing_cols)}")
            
            status_log.info("ℹ️ Transforming records and evaluating updated dependency rule matrix conditions...")
            
            # Filter rows with clean, numeric Site IDs
            alarm_df['Site ID'] = pd.to_numeric(alarm_df['Site ID'], errors='coerce')
            alarm_df = alarm_df.dropna(subset=['Site ID'])
            alarm_df['Site ID'] = alarm_df['Site ID'].astype(int)
            
            # Track max date string safely for file naming
            max_date = pd.to_datetime(alarm_df['Last Occurred On'], errors='coerce').max()
            if pd.isnull(max_date):
                max_date = datetime.datetime.now()
            
            # Build transformed arrays
            final_rows = []
            is_backbone_list = [] # Store structural flags mapping cleanly to row stylings below
            
            for _, row in alarm_df.iterrows():
                sid = row['Site ID']
                
                # Apply maps matching your exact business rules
                p_owner = str(power_co_map.get(sid, "IHS")).strip()
                is_bb = backbone_map.get(sid, "") == 'yes'
                kids_count = float(children_map.get(sid, 0))
                
                # Rule evaluation matrix block matching your fixed ternary conditional
                if is_bb:
                    prediction = f"{int(kids_count)} site" + ("s" if kids_count > 1 else "")
                elif kids_count < 5:
                    prediction = "5 sites"
                else:
                    prediction = f"{int(kids_count)} sites"
                
                final_rows.append({
                    'Site ID': sid,
                    'Site Name': row['Site Name'] if pd.notnull(row['Site Name']) else "",
                    'Power Owner': p_owner,
                    'Number of Dependencies': prediction,
                    'Number of Directly Dependent Rural Sites': "", # Safe Empty Column 5
                    'Ticket ID': row['Ticket ID'] if pd.notnull(row['Ticket ID']) else "",
                    'Alarm Name': row['Alarm Name'] if pd.notnull(row['Alarm Name']) else "",
                    'First Occurred On': str(row['First Occurred On']) if pd.notnull(row['First Occurred On']) else "",
                    'Duration(hh:mm:ss)': str(row['Duration(hh:mm:ss)']) if pd.notnull(row['Duration(hh:mm:ss)']) else "",
                    'Last Occurred On': str(row['Last Occurred On']) if pd.notnull(row['Last Occurred On']) else ""
                })
                is_backbone_list.append(is_bb)
            
            # Construct unified matrix dataframe and sort ascending by ID
            df_final = pd.DataFrame(final_rows)
            # Maintain sorting connection with the styling flag array layout 
            df_final['_is_backbone'] = is_backbone_list
            df_final = df_final.sort_values(by='Site ID').reset_index(drop=True)
            
            # Extract tracking array cleanly before printing layout rows
            sorted_backbone_flags = df_final['_is_backbone'].tolist()
            df_final = df_final.drop(columns=['_is_backbone'])
            
            status_log.info("ℹ️ Configuring presentation styles layers and applying palette styling rules...")
            
            # Step 3: Write Output Workbook Layout Matrix via openpyxl
            wb = Workbook()
            ws = wb.active
            ws.title = "Power Alarms Status"
            ws.views.sheetView[0].showGridLines = True
            
            # Match layout color tokens (Amber Fill)
            amber_fill = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")
            font_title = Font(name="Calibri", size=14, bold=True, color="000000")
            font_header = Font(name="Calibri", size=11, bold=True, color="000000")
            font_data_normal = Font(name="Calibri", size=11, color="000000")
            font_data_bold = Font(name="Calibri", size=11, bold=True, color="000000")
            
            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            
            thin_side = Side(style='thin', color='000000')
            full_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            
            # Add Title Cell
            ws.merge_cells('A1:J1') # Updated boundary to J1 to capture all 10 headers
            ws['A1'] = "Power Alarms on TIS"
            ws['A1'].font = font_title
            ws['A1'].fill = amber_fill
            ws['A1'].alignment = align_center
            ws.row_dimensions[1].height = 30
            
            # Add Headers Row
            headers = list(df_final.columns)
            ws.append(headers)
            ws.row_dimensions[2].height = 38 # Updated to match JS header height rules
            for col_idx in range(1, 11):
                cell = ws.cell(row=2, column=col_idx)
                cell.font = font_header
                cell.fill = amber_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = full_border
                
            # Append DataFrame rows matching column sequence arrays
            for r in dataframe_to_rows(df_final, index=False, header=False):
                ws.append(r)
                
            # Formatting and structural passes over all records
            for row_idx in range(3, ws.max_row + 1):
                ws.row_dimensions[row_idx].height = 18
                is_row_backbone = sorted_backbone_flags[row_idx - 3]
                
                for col_idx in range(1, 11):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = full_border
                    
                    # Apply specific Font Bold rules onto Column 4 for Backbone items
                    if col_idx == 4 and is_row_backbone:
                        cell.font = font_data_bold
                    else:
                        cell.font = font_data_normal
                        
                    # Match exact center/left column rules definitions from JavaScript
                    if col_idx in [1, 6, 8, 9, 10]:
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_left
                        
            # Dynamic Column Auto-fitting calculation loop
            for col in ws.columns:
                # Grab the column letter from the very first cell in the column safely
                col_letter = col[0].coordinate.rstrip('0123456789') 
                col_idx = col[0].column
                
                if col_idx in [4, 5]:
                    ws.column_dimensions[col_letter].width = 24
                    continue
                
                max_len = 14
                for cell in col:
                    # Skip the title and header rows to strictly measure data length values
                    if cell.row > 2 and cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                        
                ws.column_dimensions[col_letter].width = max_len + 4
                
            # Inject dynamic auto-filtering parameters boundaries
            ws.auto_filter.ref = f"A2:J{ws.max_row}"
            
            # Datetime string formatting calculations rounding to nearest 30 mins
            minutes = max_date.minute
            remainder = minutes % 30
            if remainder >= 15:
                max_date = max_date + datetime.timedelta(minutes=(30 - remainder))
            else:
                max_date = max_date - datetime.timedelta(minutes=remainder)
                
            time_string = f"{max_date.strftime('%H')}H{max_date.strftime('%M')}"
            output_filename = f"TIS_ALARMS{time_string}.xlsx"
            
            # Write to temporary memory buffer layer for Streamlit transmission download
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            status_log.success("✅ Report compiled successfully!")
            
            # Display secure Download Widget Button immediately inside UI
            st.download_button(
                label="📥 Download Processed Report Sheet",
                data=excel_buffer,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
        except Exception as err:
            status_log.error(f"❌ {str(err)}")